"""
VSS BHXH Data Collector
Module chính để thu thập dữ liệu BHXH từ website VSS
"""
import logging
import time
import json
import re
import random
from datetime import datetime, timedelta
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import requests
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side

from config import VSSConfig
from vss_authenticator import VSSAuthenticator

class VSSDataCollector:
    """Main class for collecting BHXH data from VSS website"""
    
    def __init__(self, use_proxy=True, headless=True):
        self.config = VSSConfig()
        self.authenticator = VSSAuthenticator(use_proxy=use_proxy, headless=headless)
        self.collected_data = []
        self.collection_stats = {
            "total_records": 0,
            "successful_queries": 0,
            "failed_queries": 0,
            "start_time": None,
            "end_time": None,
            "province": None
        }
        
        # Setup logging
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        
        # Ensure directories exist
        VSSConfig.ensure_directories()
        
        # Setup file handler
        log_handler = logging.FileHandler(
            VSSConfig.LOGGING_CONFIG["file"], 
            encoding='utf-8'
        )
        log_handler.setFormatter(
            logging.Formatter(VSSConfig.LOGGING_CONFIG["format"])
        )
        self.logger.addHandler(log_handler)
        
        self.logger.info("Khởi tạo VSS Data Collector")
    
    def _normalize_vietnamese_text(self, text):
        """Normalize Vietnamese text"""
        if not text or not isinstance(text, str):
            return ""
        
        # Basic text cleaning
        text = text.strip()
        text = re.sub(r'\s+', ' ', text)  # Multiple spaces to single space
        text = re.sub(r'\n+', ' ', text)  # Newlines to space
        
        # Remove extra whitespace around punctuation
        text = re.sub(r'\s*,\s*', ', ', text)
        text = re.sub(r'\s*\.\s*', '. ', text)
        
        return text
    
    def _parse_bhxh_response(self, html_content, query_params):
        """Parse BHXH lookup response HTML"""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Find result table or data container
            result_data = {}
            
            # Look for common table structures in Vietnamese government sites
            tables = soup.find_all('table')
            
            for table in tables:
                # Look for data rows
                rows = table.find_all('tr')
                
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 2:
                        # Try to extract key-value pairs
                        key = self._normalize_vietnamese_text(cells[0].get_text())
                        value = self._normalize_vietnamese_text(cells[1].get_text())
                        
                        if key and value:
                            result_data[key] = value
            
            # Look for specific BHXH fields
            bhxh_fields = {
                'ho_ten': '',
                'ngay_sinh': '',
                'gioi_tinh': '',
                'so_bhxh': '',
                'so_the_bhyt': '',
                'noi_cap': '',
                'tinh_trang': '',
                'don_vi': '',
                'dia_chi': '',
                'nghe_nghiep': ''
            }
            
            # Try to map extracted data to BHXH fields
            for field_name, default_value in bhxh_fields.items():
                # Look for field in result_data with fuzzy matching
                for key, value in result_data.items():
                    key_lower = key.lower().replace(' ', '').replace('_', '')
                    
                    if field_name == 'ho_ten' and any(x in key_lower for x in ['họtên', 'hoten', 'tên']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'ngay_sinh' and any(x in key_lower for x in ['ngaysinh', 'sinh']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'gioi_tinh' and any(x in key_lower for x in ['giới', 'phai']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'so_bhxh' and any(x in key_lower for x in ['bhxh', 'baohiem']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'so_the_bhyt' and any(x in key_lower for x in ['bhyt', 'thẻ']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'noi_cap' and any(x in key_lower for x in ['nơicấp', 'cap']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'tinh_trang' and any(x in key_lower for x in ['tinhtrang', 'trangthai']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'don_vi' and any(x in key_lower for x in ['donvi', 'cơquan']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'dia_chi' and any(x in key_lower for x in ['diachi', 'địa']):
                        bhxh_fields[field_name] = value
                    elif field_name == 'nghe_nghiep' and any(x in key_lower for x in ['nghề', 'congviec']):
                        bhxh_fields[field_name] = value
            
            # Add metadata
            bhxh_fields.update({
                'query_params': query_params,
                'collection_time': datetime.now().isoformat(),
                'source_url': self.config.BHXH_LOOKUP_URL,
                'raw_data': result_data
            })
            
            return bhxh_fields
            
        except Exception as e:
            self.logger.error("Lỗi parse BHXH response: %s", str(e))
            return {}
    
    def _perform_bhxh_lookup(self, search_params):
        """Perform BHXH lookup with given parameters"""
        try:
            self.logger.info("Thực hiện tra cứu BHXH với params: %s", search_params)
            
            # Use browser if available for better compatibility
            if self.authenticator.driver:
                return self._lookup_with_browser(search_params)
            else:
                return self._lookup_with_session(search_params)
                
        except Exception as e:
            self.logger.error("Lỗi thực hiện tra cứu BHXH: %s", str(e))
            return None
    
    def _lookup_with_browser(self, search_params):
        """Perform lookup using browser automation"""
        try:
            driver = self.authenticator.driver
            
            # Navigate to lookup page
            driver.get(self.config.BHXH_LOOKUP_URL)
            
            # Wait for page load
            wait = WebDriverWait(driver, 30)
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
            
            # Fill search form
            if 'ho_ten' in search_params:
                try:
                    name_input = wait.until(EC.presence_of_element_located((By.NAME, "hoTen")))
                    name_input.clear()
                    name_input.send_keys(search_params['ho_ten'])
                except TimeoutException:
                    # Try alternative selectors
                    name_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
                    if name_inputs:
                        name_inputs[0].clear()
                        name_inputs[0].send_keys(search_params['ho_ten'])
            
            if 'ngay_sinh' in search_params:
                try:
                    birth_input = driver.find_element(By.NAME, "ngaySinh")
                    birth_input.clear()
                    birth_input.send_keys(search_params['ngay_sinh'])
                except NoSuchElementException:
                    pass
            
            if 'so_bhxh' in search_params:
                try:
                    bhxh_input = driver.find_element(By.NAME, "soBHXH")
                    bhxh_input.clear()
                    bhxh_input.send_keys(search_params['so_bhxh'])
                except NoSuchElementException:
                    pass
            
            # Submit form
            submit_buttons = driver.find_elements(By.CSS_SELECTOR, "input[type='submit'], button[type='submit'], .btn-search")
            if submit_buttons:
                submit_buttons[0].click()
                
                # Wait for results
                time.sleep(3)
                
                # Get result page source
                result_html = driver.page_source
                return self._parse_bhxh_response(result_html, search_params)
            
            return {}
            
        except Exception as e:
            self.logger.error("Lỗi lookup với browser: %s", str(e))
            return {}
    
    def _lookup_with_session(self, search_params):
        """Perform lookup using requests session"""
        try:
            session = self.authenticator.session
            
            # Prepare form data
            form_data = {
                'hoTen': search_params.get('ho_ten', ''),
                'ngaySinh': search_params.get('ngay_sinh', ''),
                'soBHXH': search_params.get('so_bhxh', '')
            }
            
            # Add CSRF token if available
            if self.authenticator.csrf_token:
                form_data['_token'] = self.authenticator.csrf_token
            
            # Make POST request
            response = session.post(
                self.config.BHXH_LOOKUP_URL,
                data=form_data,
                timeout=VSSConfig.REQUEST_CONFIG["timeout"]
            )
            
            if response.status_code == 200:
                return self._parse_bhxh_response(response.text, search_params)
            else:
                self.logger.error("Lookup request failed with status: %d", response.status_code)
                return {}
                
        except Exception as e:
            self.logger.error("Lỗi lookup với session: %s", str(e))
            return {}
    
    def collect_province_data(self, province_info, sample_queries=None):
        """Collect BHXH data for a specific province"""
        try:
            self.collection_stats["province"] = province_info["name"]
            self.collection_stats["start_time"] = datetime.now().isoformat()
            
            self.logger.info("Bắt đầu thu thập dữ liệu cho tỉnh: %s", province_info["name"])
            
            # Authenticate first
            if not self.authenticator.authenticate():
                self.logger.error("Không thể xác thực với VSS")
                return False
            
            # Use sample queries or generate test queries
            if not sample_queries:
                sample_queries = self._generate_sample_queries(province_info)
            
            # Collect data for each query
            for i, query in enumerate(sample_queries):
                try:
                    self.logger.info("Thực hiện query %d/%d", i+1, len(sample_queries))
                    
                    # Add delay between requests
                    if i > 0:
                        delay = random.uniform(2, 5)
                        self.logger.info("Chờ %.1f giây trước query tiếp theo", delay)
                        time.sleep(delay)
                    
                    # Perform lookup
                    result = self._perform_bhxh_lookup(query)
                    
                    if result and any(result.get(field) for field in ['ho_ten', 'so_bhxh', 'so_the_bhyt']):
                        self.collected_data.append(result)
                        self.collection_stats["successful_queries"] += 1
                        self.collection_stats["total_records"] += 1
                        self.logger.info("Thu thập thành công: %s", result.get('ho_ten', 'N/A'))
                    else:
                        self.collection_stats["failed_queries"] += 1
                        self.logger.warning("Query không có kết quả: %s", query)
                        
                except Exception as e:
                    self.collection_stats["failed_queries"] += 1
                    self.logger.error("Lỗi khi thực hiện query %d: %s", i+1, str(e))
                    continue
            
            self.collection_stats["end_time"] = datetime.now().isoformat()
            self.logger.info("Hoàn thành thu thập dữ liệu. Total: %d records", 
                           self.collection_stats["total_records"])
            
            return True
            
        except Exception as e:
            self.logger.error("Lỗi thu thập dữ liệu tỉnh: %s", str(e))
            return False
    
    def _generate_sample_queries(self, province_info):
        """Generate sample queries for testing"""
        # This is a placeholder - in reality you'd want to have actual test data
        sample_queries = [
            {
                "ho_ten": "Nguyễn Văn A",
                "ngay_sinh": "01/01/1980",
                "description": f"Test query 1 for {province_info['name']}"
            },
            {
                "ho_ten": "Trần Thị B", 
                "ngay_sinh": "15/05/1985",
                "description": f"Test query 2 for {province_info['name']}"
            },
            {
                "so_bhxh": "1234567890",
                "description": f"Test query 3 for {province_info['name']}"
            }
        ]
        
        self.logger.info("Tạo %d sample queries cho %s", len(sample_queries), province_info['name'])
        return sample_queries
    
    def export_to_excel(self, output_path=None):
        """Export collected data to Excel file"""
        try:
            if not self.collected_data:
                self.logger.warning("Không có dữ liệu để export")
                return None
            
            # Generate output path if not provided
            if not output_path:
                province_name = self.collection_stats.get("province", "unknown")
                output_path = f"{VSSConfig.PATHS['output']}/{VSSConfig.get_output_filename(province_name)}"
            
            self.logger.info("Bắt đầu export dữ liệu ra Excel: %s", output_path)
            
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create summary sheet
            self._create_summary_sheet(workbook)
            
            # Create detail sheet
            self._create_detail_sheet(workbook)
            
            # Create metadata sheet
            self._create_metadata_sheet(workbook)
            
            # Save workbook
            workbook.save(output_path)
            
            self.logger.info("Export Excel thành công: %s", output_path)
            return output_path
            
        except Exception as e:
            self.logger.error("Lỗi export Excel: %s", str(e))
            return None
    
    def _create_summary_sheet(self, workbook):
        """Create summary worksheet"""
        ws = workbook.create_sheet(VSSConfig.EXCEL_CONFIG["worksheet_names"]["summary"])
        
        # Headers
        headers = [
            "STT", "Họ tên", "Ngày sinh", "Giới tính", "Số BHXH", 
            "Số thẻ BHYT", "Nơi cấp", "Tình trạng", "Thời gian thu thập"
        ]
        
        # Apply header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_idx, record in enumerate(self.collected_data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1)  # STT
            ws.cell(row=row_idx, column=2, value=record.get('ho_ten', ''))
            ws.cell(row=row_idx, column=3, value=record.get('ngay_sinh', ''))
            ws.cell(row=row_idx, column=4, value=record.get('gioi_tinh', ''))
            ws.cell(row=row_idx, column=5, value=record.get('so_bhxh', ''))
            ws.cell(row=row_idx, column=6, value=record.get('so_the_bhyt', ''))
            ws.cell(row=row_idx, column=7, value=record.get('noi_cap', ''))
            ws.cell(row=row_idx, column=8, value=record.get('tinh_trang', ''))
            ws.cell(row=row_idx, column=9, value=record.get('collection_time', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detail_sheet(self, workbook):
        """Create detailed data worksheet"""
        ws = workbook.create_sheet(VSSConfig.EXCEL_CONFIG["worksheet_names"]["detail"])
        
        if not self.collected_data:
            return
        
        # Get all possible fields from collected data
        all_fields = set()
        for record in self.collected_data:
            all_fields.update(record.keys())
        
        # Remove non-displayable fields
        exclude_fields = {'raw_data', 'query_params'}
        display_fields = sorted([f for f in all_fields if f not in exclude_fields])
        
        # Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col, field in enumerate(display_fields, 1):
            cell = ws.cell(row=1, column=col, value=field.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row_idx, record in enumerate(self.collected_data, 2):
            for col, field in enumerate(display_fields, 1):
                value = record.get(field, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col, value=str(value))
    
    def _create_metadata_sheet(self, workbook):
        """Create metadata worksheet"""
        ws = workbook.create_sheet(VSSConfig.EXCEL_CONFIG["worksheet_names"]["metadata"])
        
        metadata = [
            ("Tỉnh/Thành phố", self.collection_stats.get("province", "N/A")),
            ("Thời gian bắt đầu", self.collection_stats.get("start_time", "N/A")),
            ("Thời gian kết thúc", self.collection_stats.get("end_time", "N/A")),
            ("Tổng số records", self.collection_stats.get("total_records", 0)),
            ("Query thành công", self.collection_stats.get("successful_queries", 0)),
            ("Query thất bại", self.collection_stats.get("failed_queries", 0)),
            ("Nguồn dữ liệu", VSSConfig.BASE_URL),
            ("Phiên bản thu thập", "1.0.0"),
            ("Ghi chú", "Dữ liệu được thu thập tự động từ website VSS")
        ]
        
        for row_idx, (key, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=str(value))
    
    def save_results_json(self, output_path=None):
        """Save results to JSON file"""
        try:
            if not output_path:
                timestamp = VSSConfig.get_timestamp()
                province = self.collection_stats.get("province", "unknown").replace(" ", "_").lower()
                output_path = f"{VSSConfig.PATHS['data']}/bhxh_results_{province}_{timestamp}.json"
            
            results = {
                "collection_stats": self.collection_stats,
                "collected_data": self.collected_data,
                "config_info": {
                    "proxy_enabled": self.authenticator.use_proxy,
                    "headless_mode": self.authenticator.headless,
                    "collection_timestamp": datetime.now().isoformat()
                }
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            
            self.logger.info("Đã lưu kết quả JSON: %s", output_path)
            return output_path
            
        except Exception as e:
            self.logger.error("Lỗi lưu JSON: %s", str(e))
            return None
    
    def get_collection_summary(self):
        """Get collection summary"""
        return {
            "province": self.collection_stats.get("province"),
            "total_records": self.collection_stats.get("total_records", 0),
            "success_rate": (
                self.collection_stats.get("successful_queries", 0) / 
                max(self.collection_stats.get("successful_queries", 0) + 
                    self.collection_stats.get("failed_queries", 0), 1)
            ) * 100,
            "collection_duration": self._calculate_duration(),
            "data_fields": list(set().union(*(d.keys() for d in self.collected_data))) if self.collected_data else []
        }
    
    def _calculate_duration(self):
        """Calculate collection duration"""
        try:
            if self.collection_stats.get("start_time") and self.collection_stats.get("end_time"):
                start = datetime.fromisoformat(self.collection_stats["start_time"])
                end = datetime.fromisoformat(self.collection_stats["end_time"])
                return str(end - start)
        except:
            pass
        return "N/A"
    
    def close(self):
        """Clean up resources"""
        try:
            self.authenticator.close()
            self.logger.info("Đã đóng VSS Data Collector")
        except Exception as e:
            self.logger.error("Lỗi khi đóng collector: %s", str(e))

# Test collector if run directly
if __name__ == "__main__":
    # Setup console logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler()]
    )
    
    # Test with a sample province
    test_province = {"name": "Hà Nội", "code": "01"}
    
    collector = VSSDataCollector(use_proxy=True, headless=True)
    
    try:
        # Collect sample data
        if collector.collect_province_data(test_province):
            print("✅ Thu thập dữ liệu thành công")
            
            # Export to Excel
            excel_path = collector.export_to_excel()
            if excel_path:
                print(f"✅ Export Excel thành công: {excel_path}")
            
            # Save JSON
            json_path = collector.save_results_json()
            if json_path:
                print(f"✅ Lưu JSON thành công: {json_path}")
            
            # Show summary
            summary = collector.get_collection_summary()
            print(f"📊 Summary: {json.dumps(summary, ensure_ascii=False, indent=2)}")
            
        else:
            print("❌ Thu thập dữ liệu thất bại")
            
    finally:
        collector.close()
